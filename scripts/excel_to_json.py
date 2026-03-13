"""
Convierte el Excel de informe semanal a public/data.json
Uso: python scripts/excel_to_json.py data/informe.xlsx public/data.json
"""
import json
import sys
from collections import defaultdict
import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_sheet(wb, candidates):
    for name in wb.sheetnames:
        for c in candidates:
            if c.upper() in name.upper():
                return wb[name]
    return None


def str_val(v):
    if v is None:
        return " "
    s = str(v).strip()
    return s if s else " "


def int_val(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def float_val(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_crecimiento(v):
    """
    Excel almacena porcentajes como fracción decimal (7.57% → 0.0757).
    Queremos el valor de display: 7.57.
    """
    if v is None:
        return 0.0
    if isinstance(v, str):
        cleaned = v.strip().replace("%", "").replace(",", ".").strip()
        try:
            return round(float(cleaned), 4)
        except ValueError:
            return 0.0
    try:
        f = float(v)
        if -1.0 < f < 1.0:
            return round(f * 100, 4)
        return round(f, 4)
    except (TypeError, ValueError):
        return 0.0


def format_date(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


# ---------------------------------------------------------------------------
# Hoja 1 — CONFIG
# ---------------------------------------------------------------------------

def parse_config(ws):
    result = {"REPORT_DATE_LABEL": "", "REPORT_DATE_FOOTER": ""}
    for row in ws.iter_rows(values_only=True):
        if not row[0]:
            continue
        campo = str(row[0]).lower()
        valor = str(row[1]).strip() if row[1] is not None else ""
        if "etiqueta" in campo:
            result["REPORT_DATE_LABEL"] = valor
        elif "pie" in campo:
            result["REPORT_DATE_FOOTER"] = valor
    return result


# ---------------------------------------------------------------------------
# Hoja 2 — CONSEJOS
# ---------------------------------------------------------------------------

CONSEJO_COL_MAP = {
    "CONSEJO":                    ("consejo",               "str"),
    "VALIDADOS":                  ("validados",             "int"),
    "OBJETIVO":                   ("objetivo",              "int"),
    "CUPO":                       ("cupo",                  "float"),
    "CRECIMIENTO %":              ("crecimiento",           "pct"),
    "CRECIMIENTO":                ("crecimiento",           "pct"),
    "INACTIVOS":                  ("inactivos",             "int"),
    "CURSOS":                     ("cursos",                "int"),
    "MAT CONSEJ":                 ("mat_consejo",           "int"),
    "MAT CONSEJO":                ("mat_consejo",           "int"),
    "MAT LO":                     ("mat_lo",                "int"),
    "MAT PROD":                   ("mat_prod",              "int"),
    "CONTRATOS (S/TOTAL)":        ("contratos",             "str"),
    "CONTRATOS (N/TOTAL)":        ("contratos",             "str"),
    "CONTRATOS":                  ("contratos",             "str"),
    "DACI (S/TOTAL)":             ("daci",                  "str"),
    "DACI (N/TOTAL)":             ("daci",                  "str"),
    "DACI":                       ("daci",                  "str"),
    "FACTURAS (S/TOTAL)":         ("facturas",              "str"),
    "FACTURAS (N/TOTAL)":         ("facturas",              "str"),
    "FACTURAS":                   ("facturas",              "str"),
    "WEB CONV":                   ("web_conv",              "int"),
    "WEB REAL":                   ("web_real",              "int"),
    "PRES CONV":                  ("pres_conv",             "int"),
    "PRES REAL":                  ("pres_real",             "int"),
    "BLOQUEOS":                   ("bloqueos",              "str"),
    "PEND. WEBINAR":              ("pendientes_webinar",    "str"),
    "PENDIENTES WEBINAR":         ("pendientes_webinar",    "str"),
    "PEND. PRESENCIAL":           ("pendientes_presencial", "str"),
    "PENDIENTES PRESENCIAL":      ("pendientes_presencial", "str"),
    "CONSULTAS / INCIDENCIAS":    ("consultas",             "str"),
    "CONSULTAS/INCIDENCIAS":      ("consultas",             "str"),
    "CONSULTAS":                  ("consultas",             "str"),
}


def parse_consejos(ws):
    headers = None
    items = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            if row[0] and str(row[0]).strip().upper() == "CONSEJO":
                headers = [str(h).strip().upper() if h else "" for h in row]
            continue
        if not row[0]:
            continue
        item = {}
        for i, header in enumerate(headers):
            if i >= len(row):
                break
            mapping = CONSEJO_COL_MAP.get(header)
            if mapping is None:
                continue
            key, dtype = mapping
            val = row[i]
            if dtype == "int":
                item[key] = int_val(val)
            elif dtype == "float":
                item[key] = float_val(val)
            elif dtype == "pct":
                item[key] = parse_crecimiento(val)
            else:
                item[key] = str_val(val)
        if item.get("consejo"):
            items.append(item)
    return items


# ---------------------------------------------------------------------------
# Hoja 3 — EVIDENCIAS
# ---------------------------------------------------------------------------

def parse_evidencias(ws):
    by_consejo = defaultdict(lambda: defaultdict(list))
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] and str(row[0]).strip().upper() == "CONSEJO":
                header_found = True
            continue
        if not row[0]:
            continue
        consejo   = str(row[0]).strip()
        sesion    = str(row[1]).strip() if row[1] else ""
        evidencia = str(row[2]).strip() if row[2] else ""
        if sesion and evidencia:
            by_consejo[consejo][sesion].append(evidencia)
    return by_consejo


def attach_evidencias(initial_data, evidencias_by_consejo):
    for item in initial_data:
        nombre = item.get("consejo", "")
        if nombre in evidencias_by_consejo:
            item["evidencias_faltantes"] = [
                {"sesion": s, "evidencias": evs}
                for s, evs in evidencias_by_consejo[nombre].items()
            ]
    return initial_data


# ---------------------------------------------------------------------------
# Hoja 4 — EVOLUTIVOS
# ---------------------------------------------------------------------------

def parse_evolutivos(ws):
    plataforma, sug = [], []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] and str(row[0]).strip().upper() in ("SECCIÓN", "SECCION", "SECTION"):
                header_found = True
            continue
        if not row[0] or not row[2]:
            continue
        seccion = str(row[0]).strip().upper()
        desc    = str(row[2]).strip()
        if "SOPORTE" in seccion or "PLATAFORMA" in seccion:
            plataforma.append(desc)
        elif "SUG" in seccion:
            sug.append(desc)
    return {"plataforma": plataforma, "sug": sug}


# ---------------------------------------------------------------------------
# Hoja 5 — HISTÓRICO GLOBAL
# ---------------------------------------------------------------------------

def parse_global_historical(ws):
    records = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] and "fecha" in str(row[0]).lower():
                header_found = True
            continue
        if row[0] is None:
            continue
        date_str = format_date(row[0])
        if not date_str:
            continue
        try:
            total = int(float(row[1]))
            records.append({"date": date_str, "Total": total})
        except (TypeError, ValueError):
            pass
    return records


# ---------------------------------------------------------------------------
# Hoja 6 — HISTÓRICO CONSEJOS
# ---------------------------------------------------------------------------

def parse_council_historical(ws):
    result = {}
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            if row[0] and "fecha" in str(row[0]).lower():
                headers = [str(h).strip() if h else "" for h in row]
                for h in headers[1:]:
                    if h:
                        result[h] = []
            continue
        if not row[0]:
            continue
        date_str = format_date(row[0])
        if not date_str:
            continue
        for i, header in enumerate(headers[1:], start=1):
            if not header or i >= len(row):
                continue
            try:
                total = int(float(row[i])) if row[i] is not None else 0
            except (TypeError, ValueError):
                total = 0
            result.setdefault(header, []).append({"date": date_str, "Total": total})
    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def excel_to_json(excel_path, output_path):
    print(f"Leyendo {excel_path} ...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"Hojas encontradas: {wb.sheetnames}")

    ws_config = find_sheet(wb, ["CONFIG"]) or wb.worksheets[0]
    config = parse_config(ws_config)

    ws_consejos = find_sheet(wb, ["CONSEJO"]) or wb.worksheets[1]
    initial_data = parse_consejos(ws_consejos)
    print(f"  Consejos encontrados: {len(initial_data)}")

    ws_evidencias = find_sheet(wb, ["EVIDENCIA"]) or wb.worksheets[2]
    evidencias = parse_evidencias(ws_evidencias)
    initial_data = attach_evidencias(initial_data, evidencias)

    ws_evolutivos = find_sheet(wb, ["EVOLUTIVO"]) or wb.worksheets[3]
    evolutivos = parse_evolutivos(ws_evolutivos)

    ws_global = find_sheet(wb, ["GLOBAL"]) or wb.worksheets[4]
    global_hist = parse_global_historical(ws_global)
    print(f"  Registros histórico global: {len(global_hist)}")

    ws_hist_consejos = find_sheet(wb, ["HISTÓRICO CONSEJO", "HISTORICO CONSEJO"]) or wb.worksheets[5]
    council_hist = parse_council_historical(ws_hist_consejos)
    print(f"  Consejos en histórico: {len(council_hist)}")

    result = {
        **config,
        "INITIAL_DATA": initial_data,
        "COUNCIL_HISTORICAL_DATA": council_hist,
        "GLOBAL_HISTORICAL": global_hist,
        "EVOLUTIVOS_DATA": evolutivos,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"✅ {output_path} generado correctamente.")


if __name__ == "__main__":
    excel_in = sys.argv[1] if len(sys.argv) > 1 else "data/informe.xlsx"
    json_out = sys.argv[2] if len(sys.argv) > 2 else "public/data.json"
    excel_to_json(excel_in, json_out)
